#CÃ¢u 11: Xá»­ lÃ½ Excel File - Viáº¿t pháº§n má»m Quáº£n LÃ½ NhÃ¢n ViÃªn

'''
YÃªu cáº§u:
Viáº¿t pháº§n má»m quáº£n lÃ½ NhÃ¢n viÃªn lÆ°u báº±ng Excel. Má»—i nhÃ¢n viÃªn cÃ³ MÃ£, TÃªn, Tuá»•i.
âˆ’ Pháº§n má»m cho phÃ©p lÆ°u NhÃ¢n viÃªn vÃ o File Excel
âˆ’ Pháº§n má»m cho phÃ©p Ä‘á»c danh sÃ¡ch NhÃ¢n viÃªn trong File Excel
âˆ’ Pháº§n má»m cho phÃ©p sáº¯p xáº¿p NhÃ¢n viÃªn theo Tuá»•i tÄƒng dáº§n

'''
from openpyxl import Workbook, load_workbook

# ======= HÃ€M GHI DANH SÃCH NHÃ‚N VIÃŠN VÃ€O FILE EXCEL =======
def ghi_nhan_vien(file_path, danh_sach_nv):
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"

    # Ghi tiÃªu Ä‘á» cá»™t
    ws.append(["STT", "MÃ£", "TÃªn", "Tuá»•i"])

    # Ghi dá»¯ liá»‡u nhÃ¢n viÃªn vÃ o báº£ng
    for i, nv in enumerate(danh_sach_nv, start=1):
        ws.append([i, nv["ma"], nv["ten"], nv["tuoi"]])

    wb.save(file_path)
    print(f"âœ… ÄÃ£ lÆ°u {len(danh_sach_nv)} nhÃ¢n viÃªn vÃ o file '{file_path}'.")


# ======= HÃ€M Äá»ŒC DANH SÃCH NHÃ‚N VIÃŠN Tá»ª FILE EXCEL =======
def doc_nhan_vien(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    danh_sach_nv = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Bá» hÃ ng tiÃªu Ä‘á»
        nv = {"stt": row[0], "ma": row[1], "ten": row[2], "tuoi": row[3]}
        danh_sach_nv.append(nv)

    return danh_sach_nv


# ======= HÃ€M Sáº®P Xáº¾P NHÃ‚N VIÃŠN THEO TUá»”I TÄ‚NG Dáº¦N =======
def sap_xep_theo_tuoi(danh_sach_nv):
    return sorted(danh_sach_nv, key=lambda x: x["tuoi"])


# ======= CHÆ¯Æ NG TRÃŒNH CHÃNH =======
file_excel = "NhanVien.xlsx"

# 1ï¸âƒ£ Nháº­p danh sÃ¡ch nhÃ¢n viÃªn máº«u
danh_sach_nv = [
    {"ma": "NV1", "ten": "An", "tuoi": 18},
    {"ma": "NV2", "ten": "LÃ nh", "tuoi": 22},
    {"ma": "NV3", "ten": "Giáº£i", "tuoi": 20},
    {"ma": "NV4", "ten": "ThoÃ¡t", "tuoi": 19},
    {"ma": "NV5", "ten": "Háº¡nh", "tuoi": 25},
    {"ma": "NV6", "ten": "PhÃºc", "tuoi": 24},
]

# 2ï¸âƒ£ Ghi danh sÃ¡ch vÃ o file Excel
ghi_nhan_vien(file_excel, danh_sach_nv)

# 3ï¸âƒ£ Äá»c láº¡i danh sÃ¡ch tá»« file Excel
ds_doc = doc_nhan_vien(file_excel)
print("\nğŸ“‹ Danh sÃ¡ch nhÃ¢n viÃªn Ä‘á»c tá»« file:")
for nv in ds_doc:
    print(nv)

# 4ï¸âƒ£ Sáº¯p xáº¿p danh sÃ¡ch theo tuá»•i tÄƒng dáº§n
ds_sap_xep = sap_xep_theo_tuoi(ds_doc)
print("\nğŸ“ˆ Danh sÃ¡ch nhÃ¢n viÃªn sau khi sáº¯p xáº¿p theo tuá»•i tÄƒng dáº§n:")
for nv in ds_sap_xep:
    print(nv)
