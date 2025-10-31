#Câu 8: Xử lý đọc Excel File
'''
Yêu cầu:
Sử dụng thư viện openpyxl để đọc file excel ở câu trước
'''


# Đọc file Excel bằng openpyxl

from openpyxl import load_workbook

# Mở file Excel
workbook = load_workbook("sanpham.xlsx")

# Chọn sheet đầu tiên
sheet = workbook.active

print(" DỮ LIỆU TRONG FILE EXCEL:\n")

# Duyệt từng dòng dữ liệu
for row in sheet.iter_rows(values_only=True):
    print(row)

# Đóng file Excel
workbook.close()
